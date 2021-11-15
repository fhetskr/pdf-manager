import os
import pdfrw
import openpyxl
import json

### Internal data types
### Credits:
### Peter: GenericFile
### Matthew: PdfFile, CsvFile, TxtFile, small part of GenericFile.convert(), GenericFile.convert_redirect()
### Rohit: ExcelFile, JsonFile

class OrderedDictionary():
	# NB: I'm not sure we actually need to implement this -Peter
	# Edit: nvm I think we probably will have to
	pass


class FileError(Exception):
	'''Exception for invalid files'''
	pass

class InvalidKeyPairFormat(Exception):
	# Exception when there are more than a key and value pair format
	# within a file
	pass

class KeyRepetition(Exception):
	# Exception when there are multiple of the same key
	pass

class NonexistentPdfFile(Exception):
	# Exception when a PDF being filled does not exist
	pass

# A placeholder file type
class GenericFile():
	
	def __init__(self, path):
		'''Generic constructor for file types'''
		'''By Peter'''
		# save path to self.path
		self.path = path
		self.contents = {}
		# get file extension
		if len(path.split('.')) > 1:
			self.extension = path.split('.')[-1]
		
		if os.path.isfile(path):
			if not self.verify():
				raise FileError('Malformed input')
			self.read()
		else:
			# create an empty file if none exists.
			with open(path, 'w') as f:
				pass
	
	def read(self):
		'''Updates self.contents with the actual data from the file at self.path'''
		pass
		
	def write(self):
		'''Writes contents to the file at self.path'''
		pass
	
	def convert(self, target_type, target_path):
		'''Return a file object of target_type with this file's contents'''
		'''By Peter, with some adjustments by Matthew'''
		if target_path:
			path = target_path
		else:
			# strip old extension if necessary
			if file_dict[self.path.split('.')[-1]] == self.__class__.__name__:
				path = '.'.join(self.path.split('.')[:-1])
			else:
				path = self.path
			# add new extension
			path = '.'.join((path,target_type))
		# create new file object and return it
		new_file = file_dict[target_type](path)
		new_file.contents = self.contents
		return new_file

	def convert_redirect(self, target_type, target_path, redir_file):
		if target_path:
			path = target_path
		else:
			# strip old extension if necessary
			if file_dict[self.path.split('.')[-1]] == self.__class__.__name__:
				path = '.'.join(self.path.split('.')[:-1])
			else:
				path = self.path
			# add new extension
			path = '.'.join((path,target_type))
		# create new file object and return it
		new_file = file_dict[target_type](path)
		redir_file.read_keys()
		# transpose the keys, but only if the two files have equal key count
		if len(self.contents.keys()) != len(redir_file.contents.keys()):
			raise FileError("Key alias files must have exactly as many keys as the file being aliased.")
		for i in range(0,len(self.contents.keys())):
			new_file.contents[list(redir_file.contents)[i]] = self.contents[list(self.contents)[i]]
		return new_file
		
	def verify(self):
		'''Verify that self.path points to a file of the appropriate format'''
                # Ensure that filetypes that do not implement this method automatically pass (i.e. TxtFile for quick testing purposes).
		return True

	def delete(self):
		'''Deletes the file at self.path from the filesystem, but retains contents'''
		os.remove(self.path)


### Actual file types

#Note: When doing error handling, make sure you write an exception that handles when a user has
# the Excel file open while trying to read from it.
class ExcelFile(GenericFile):
	# Reads data from a Excel file and sets the key value pairs into the contents variable
	# By Rohit Ramakrishnan
	def read(self):
		workbook = openpyxl.load_workbook(self.path)
		sheet = workbook.active

		for i in range(1, sheet.max_row + 1):
			for j in range(1, sheet.max_column + 1):
				cell_row_data = sheet.cell(row=i, column=j)
				cell_col_data = sheet.cell(row=i, column=j + 1)
				self.contents[cell_row_data.value] = cell_col_data.value
				break

	# Takes data from the contents variable and writes it into a Excel file
	# By Rohit Ramakrishnan
	def write(self):
		workbook = openpyxl.Workbook()
		sheet = workbook.active
		row_traversal = 1

		for key, value in self.contents.items():
			for j in range(1, 3):
				cell_data = sheet.cell(row=row_traversal, column=j)
				if j == 1:
					cell_data.value = key
				if j == 2:
					cell_data.value = value
			row_traversal += 1

		workbook.save(self.path)

	def verify(self):
		workbook = openpyxl.load_workbook(self.path)
		sheet = workbook.active

		if sheet.max_column != 2:
			raise InvalidKeyPairFormat
		else:
			key_list = []
			for i in range(1, sheet.max_row + 1):
				cell_data = sheet.cell(row=i, column=1)
				if cell_data.value in key_list:
					raise KeyRepetition
				else:
					key_list.append(cell_data.value)
		return True

class TxtFile(GenericFile):
	def read_keys(self):
		f = open(self.path, 'r')
		self.contents = {}
		lines = f.read().splitlines()
		for line in lines:
			self.contents[line] = None
		f.close()
	
	def write_keys(self):
		f = open(self.path, 'w')
		ret = '\n'.join(list(self.contents))
		f.write(ret)
		f.close()


class CsvFile(GenericFile):
	def read(self):
		'''Updates self.contents with the actual data from the file at self.path'''
		'''By Matthew'''
		f = open(self.path, 'r')
		self.contents = {}
		lines = f.readlines()
		for line in lines:
			data = None
			quotes = False
			# handle values wrapped in quote marks (for commas in values)
			if '","' in line:
				data = line.split('","')
				quotes = True
			else:
				data = line.split(',')
			if quotes:
				# trim any remaining quote marks
				self.contents[data[0][1:]] = data[1][:-2]
			else:
				self.contents[data[0]] = data[1]
		f.close()

	def write(self):
		'''Writes contents to the file at self.path'''
		'''By Matthew'''
		f = open(self.path, 'w')
		ret = ''
		for key, value in self.contents.items():
			ret += '"%s","%s"\n' % (key, value)
		f.write(ret)
		f.close()

	def verify(self):
		lines = open(self.path, 'r').read().splitlines()
		keys = []
		for line in lines:
			split = None
			if '","' in line:
				split = line.split('","')
			else:
				split = line.split(',')
			if len(split) != 2:
				raise InvalidKeyPairFormat
			if split[0] in keys:
				raise KeyRepetition
			keys.append(split[0])
		return True



class PdfFile(GenericFile):
	def read(self):
		'''Updates self.contents with the actual data from the file at self.path'''
		'''Adapted from https://akdux.com/python/2020/10/31/python-fill-pdf-files.html'''
		# store the entire pdf in data
		self.data = pdfrw.PdfReader(self.path)
		self.contents = {}
		# iterate over pages
		for page in self.data.pages:
			annots = page['/Annots']
			# then over annotations on that page
			for annot in annots:
				# check that we're in a widget with a field key
				if annot['/Subtype'] == '/Widget':
					if annot['/T']:
						key = annot['/T'][1:-1]
						val = ''
						# read the field value, if any
						if annot['/AS']:
							if annot['/AS'] == '/Yes':
								val = True
							if annot['/AS'] == '/Off':
								val = False
						elif annot['/V']:
							val = annot['/V'][1:-1]
						self.contents[key] = val
				# ugly handling of edge cases in case a field is embedded entirely within the parent value
				if annot['/Parent']:
					if annot['/Parent']['/T']:
						key = annot['/Parent']['/T'][1:-1]
						val = ''
						# read the field value, if any
						if annot['/AS']:
							if annot['/AS'] == '/Yes':
								val = True
							if annot['/AS'] == '/Off':
								val = False
						elif annot['/V']:
							val = annot['/V'][1:-1]
						self.contents[key] = val
	
	def write(self):
		'''Writes contents to the file at self.path'''
		'''Adapted from https://akdux.com/python/2020/10/31/python-fill-pdf-files.html'''
		# sanity check: data exists if read occurred
		if self.data is None:
			raise NonexistentPdfFile
		# first, flush any changes in contents to data
		# iterate over pages
		for page in self.data.pages:
			annots = page['/Annots']
			# then over annotations on that page
			for annot in annots:
				# check that we're in a widget with a field key
				if annot['/Subtype'] == '/Widget':
					if annot['/T']:
						key = annot['/T'][1:-1]
						if key in self.contents.keys():
							if type(self.contents[key]) == bool or self.contents[key] == 'True' or self.contents[key] == 'False':
								# special case for boolean values (checkboxes, radio buttons)
								if self.contents[key] or self.contents[key] == 'True':
									annot.update(pdfrw.PdfDict(AS=pdfrw.PdfName('Yes')))
								else:
									annot.update(pdfrw.PdfDict(AS=pdfrw.PdfName('Off')))
							else:
								annot.update(pdfrw.PdfDict(V='{}'.format(self.contents[key])))
								annot.update(pdfrw.PdfDict(AP=''))
		# trick acrobat into thinking the fields are populated
		self.data.Root.AcroForm.update(pdfrw.PdfDict(NeedAppearances=pdfrw.PdfObject('true')))
		# and finally actually write
		pdfrw.PdfWriter().write(self.path, self.data)

	def write_ordered(self):
		'''Version of write that writes values strictly in the order they appear'''
		i = 0
		if self.data is None:
			raise NonexistentPdfFile
		# similar flow to write()
		for page in self.data.pages:
			annots = page['/Annots']
			for annot in annots:
				if annot['/Subtype'] == '/Widget':
					if annot['/T']:
						# logic changes here to account for order-based writing
						if type(self.contents.values[i]) == bool:
							if self.contents.values[i]:
								annot.update(pdfrw.PdfDict(AS=pdfrw.PdfName('Yes')))
							else:
								annot.update(pdfrw.PdfDict(AS=pdfrw.PdfName('Off')))
						else:
							annot.update(pdfrw.PdfDict(V='{}'.format(self.contents.values[i])))
							annot.update(pdfrw.PdfDict(AP=''))
		self.data.Root.AcroForm.update(pdfrw.PdfDict(NeedAppearances=pdfrw.PdfObject('true')))
		pdfrw.PdfWriter().write(self.path, self.data)


class JsonFile(GenericFile):
	# Reads data from a json file and sets the key value pairs into the contents variable
	# By Rohit Ramakrishnan
	def read(self):
		with open(self.path) as f:
			self.contents = json.load(f)

	# Take data from the contents variable and writes it into a json file
	# By Rohit Ramakrishnan
	def write(self):
		# Writes contents to the file at self.path
		# By Rohit Ramakrishnan
		with open(self.path, 'w') as json_file:
			json.dump(self.contents, json_file)


'''Dictionary for file extensions'''
file_dict = {
	"txt": TxtFile,
	"csv": CsvFile,
	"xlsx": ExcelFile,
	"pdf": PdfFile,
	"json": JsonFile
}
